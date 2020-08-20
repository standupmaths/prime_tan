def tan(start=0):
    # finding values for n for which tan(n) >= n
    # this version allows neg values of tan to count
    
    # it actual searches for intergers near the multiples of pi
    # so it gets 2 instead of 1 at the start (ignore that)

    import math
    import openpyxl
    
    # going to save anything I find in a spreadsheet
    filename = "/Users/mattparker/Desktop/tan.xlsx"
    
    wb = openpyxl.load_workbook(filename, data_only=True)
    # that data_only=True bit means I get the values not the formulas
    # I leave it on by default
    
    data = wb[u'Sheet1']
    
    active_row = 1
    # move down to the first empty row in the spreadsheet
    while str(data.cell(row=active_row, column=1).value) != 'None':
        active_row += 1
    
    pi = math.pi
    
    offset = int((start-pi)/pi)
    
    value = pi/2 + pi*offset
    
    while value <= 10**46:  # go big or go home
        
        this_int = round(value,0)
        if abs(math.tan(this_int)) >= this_int:
            print this_int
            data.cell(row=active_row, column=1).value = this_int
            data.cell(row=active_row, column=2).value = math.tan(this_int)
            wb.save(filename)
            active_row += 1
        value += pi
        
    
    return 'DONE'
   